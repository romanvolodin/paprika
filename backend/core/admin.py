import subprocess
from datetime import datetime
from pathlib import Path

from django.contrib import admin
from django.core.exceptions import ObjectDoesNotExist
from django.core.files import File
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.utils.safestring import mark_safe

from .forms import (
    AddShotsToGroupsForm,
    AddShotsToProjectForm,
    AddTasksToShotForm,
    AddСommentToShotForm,
)
from .models import (
    Attachment,
    ChatMessage,
    Project,
    Shot,
    ShotGroup,
    ShotTask,
    Status,
    Task,
    TmpShotPreview,
    Version,
)


admin.site.site_header = "Паприка"
admin.site.site_title = "Паприка"
admin.site.index_title = "Администрирование Паприки"


@admin.register(Project)
class ProjectAdmin(admin.ModelAdmin):
    save_on_top = True
    readonly_fields = (
        "created_by",
        "created_at",
    )
    list_display = (
        "name",
        "code",
        "created_at",
    )

    def save_model(self, request, obj, form, change):
        obj.created_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(ShotGroup)
class ShotGroupAdmin(admin.ModelAdmin):
    save_on_top = True
    ordering = ("-name",)
    readonly_fields = (
        "created_by",
        "created_at",
    )

    def save_model(self, request, obj, form, change):
        obj.created_by = request.user
        super().save_model(request, obj, form, change)


class VersionInline(admin.TabularInline):
    model = Version
    show_change_link = True
    extra = 0


class ChatMessageInline(admin.TabularInline):
    model = ChatMessage
    show_change_link = True
    extra = 0
    fields = (
        "created_by",
        "created_at",
        "reply_to",
        "text",
    )


class ShotTaskInline(admin.TabularInline):
    model = ShotTask
    show_change_link = True
    extra = 0


class ShotStatusListFilter(admin.SimpleListFilter):
    title = "shot status"
    parameter_name = "shot_status"

    def lookups(self, request, model_admin):
        return [
            ("no_tasks", "Нет задач"),
            ("canceled", "Отмена"),
            ("not_started", "Не начат"),
            ("in_progress", "В работе"),
            ("done", "Готов"),
            ("commented", "Есть комментарий"),
            ("approved", "Принят"),
            ("delivered", "Отдан"),
        ]

    def queryset(self, request, queryset):
        if self.value() == "no_tasks":
            return queryset.filter(task_statuses=None)

        if self.value() == "canceled":
            return queryset.filter(task_statuses__status__title="Отмена")

        if self.value() == "not_started":
            return queryset.filter(task_statuses__status__title="Не начата")

        if self.value() == "in_progress":
            return queryset.filter(task_statuses__status__title="В работе")

        if self.value() == "done":
            return queryset.filter(task_statuses__status__title="Готова")

        if self.value() == "commented":
            return queryset.filter(task_statuses__status__title="Есть комментарий")

        if self.value() == "approved":
            return queryset.filter(task_statuses__status__title="Принята")

        if self.value() == "delivered":
            return queryset.filter(task_statuses__status__title="Отдано")


@admin.register(Shot)
class ShotAdmin(admin.ModelAdmin):
    save_on_top = True
    list_display = (
        "get_version_preview",
        "get_name",
        "get_rec_timecode",
        "get_shot_status",
        "get_shot_tasks",
        "get_assigned_to",
        "get_chat_message",
        "get_shot_groups",
    )
    list_filter = (
        "group",
        ShotStatusListFilter,
        "task",
    )
    inlines = (
        ChatMessageInline,
        ShotTaskInline,
        VersionInline,
    )
    ordering = ("name",)
    readonly_fields = (
        "created_by",
        "created_at",
    )

    fieldsets = (
        (
            None,
            {
                "fields": (
                    ("name", "group", "project"),
                    ("rec_timecode", "scene"),
                )
            },
        ),
        (
            "Информация об исходнике",
            {
                "fields": (
                    (
                        "source_name",
                        "source_start_timecode",
                        "source_end_timecode",
                    ),
                )
            },
        ),
        (
            "Дополнительная информация",
            {
                "fields": (
                    (
                        "pixel_aspect",
                        "retime_speed",
                        "created_by",
                        "created_at",
                    ),
                    ("comment",),
                )
            },
        ),
    )

    @admin.display(description="Шот")
    def get_name(self, shot):
        return mark_safe(f'<a href="/{shot.project.code}/shots/{shot.name}">{shot.name}</a>')

    @admin.display(description="TC")
    def get_rec_timecode(self, obj):
        return obj.rec_timecode

    @admin.display(description="TC исходника")
    def get_source_timecode(self, obj):
        return obj.source_start_timecode

    @admin.display(description="Latest version")
    def get_latest_version(self, obj):
        latest_version = obj.versions.latest("created_at")
        if latest_version:
            return mark_safe(f'<a href="{latest_version.video.url}">{latest_version.name}</a>')

    @admin.display(description="Комментарий")
    def get_chat_message(self, obj):
        return obj.chat_messages.latest("created_at").text

    @admin.display(description="группы")
    def get_shot_groups(self, obj):
        return mark_safe("<br>".join([group.name for group in obj.group.all()]))

    @admin.display(description="задачи")
    def get_shot_tasks(self, obj):
        return mark_safe("<br>".join([task.description for task in obj.task.all()]))

    @admin.display(description="превью")
    def get_version_preview(self, shot):
        try:
            version = shot.versions.latest()
        except ObjectDoesNotExist:
            return mark_safe(
                "<div style='"
                "aspect-ratio:2.39/0.2;"
                "width:150px;"
                "background-color:#eee;"
                "display:flex;"
                "align-items:center;"
                "justify-content:center;"
                "' "
                ">Нет превью</div>"
            )

        return mark_safe(
            f"<img src='{version.preview.url}' style='aspect-ratio:2.39/1; width:150px; object-fit:cover'>"
        )

    @admin.display(description="Shot status")
    def get_shot_status(self, obj):
        statuses = [
            shot_task.status.title
            for shot_task in obj.task_statuses.all()
            if shot_task.task.description != "Выдать материал"
        ]
        template = "<span style='background-color:{};color:#fff;padding:3px 7px'>{}</span>"

        if set() == set(statuses):
            return mark_safe(template.format("#ccc", "Нет задач"))

        if set(("Отмена",)) == set(statuses):
            return mark_safe(template.format("#999", "Отмена"))

        if set(("На паузе",)) == set(statuses) or set(("Отмена", "На паузе")) == set(statuses):
            return mark_safe(template.format("#3d3d3d", "На паузе"))

        if (
            set(("Не начата",)) == set(statuses)
            or set(("Не начата", "На паузе")) == set(statuses)
            or set(("Отмена", "Не начата", "На паузе")) == set(statuses)
        ):
            return mark_safe(template.format("#900", "Не начат"))

        if set(("Принята",)) == set(statuses) or set(("Отмена", "Принята", "На паузе")) == set(
            statuses
        ):
            return mark_safe(template.format("#099", "Принят"))

        if set(("Готова",)) == set(statuses) or set(("Отмена", "Готова", "На паузе")) == set(
            statuses
        ):
            return mark_safe(template.format("#090", "Готов"))

        if set(("Отдано",)) == set(statuses) or set(("Отмена", "Отдано", "На паузе")) == set(
            statuses
        ):
            return mark_safe(template.format("#7700BE", "Отдан"))

        if "Есть комментарий" in statuses:
            return mark_safe(template.format("#FF6600", "Есть комментарий"))

        return mark_safe(template.format("#fa0", "В работе"))

    @admin.display(description="Назначено")
    def get_assigned_to(self, shot):
        assignees = [task.assigned_to for task in shot.task_statuses.all() if task.assigned_to]
        html = []
        for assignee in assignees:
            if assignee.avatar:
                html.append(
                    f"<img src='{assignee.avatar.url}' style='width:30px;height:30px;border-radius:50vh'>"
                )
            else:
                html.append(
                    f"<div style='display:flex;align-items:center;justify-content:center;background-color:lightgray;width:30px;height:30px;border-radius:50vh'>{assignee.first_name[0]}</div>"
                )

        return mark_safe(" ".join(html))

    actions = [
        "add_shots_to_project",
        "add_shots_to_groups",
        "remove_shots_from_groups",
        "add_tasks_to_shot",
        "add_comments_to_shot",
        "download_shots_as_xlsx",
    ]

    @admin.action(description="Добавить в проект")
    def add_shots_to_project(modeladmin, request, queryset):
        if "apply" in request.POST:
            form = AddShotsToProjectForm(request.POST)
            if form.is_valid():
                project = form.cleaned_data["project"]
                for shot in queryset:
                    shot.project = project
                    shot.save()

            return HttpResponseRedirect(request.get_full_path())

        context = {
            "shots": queryset,
            "form": AddShotsToProjectForm,
        }
        return render(request, "admin/add_shots_to_project.html", context=context)

    @admin.action(description="Добавить задачи")
    def add_tasks_to_shot(modeladmin, request, queryset):
        if "apply" in request.POST:
            form = AddTasksToShotForm(request.POST)
            if form.is_valid():
                tasks = form.cleaned_data["tasks"]
                status = Status.objects.get(title="Не начата")
                for shot in queryset:
                    for task in tasks:
                        try:
                            ShotTask.objects.get(
                                shot=shot,
                                task=task,
                                status=status,
                            )
                            continue
                        except ObjectDoesNotExist:
                            ShotTask.objects.create(
                                shot=shot,
                                task=task,
                                status=status,
                            )

            return HttpResponseRedirect(request.get_full_path())

        context = {
            "shots": queryset,
            "form": AddTasksToShotForm,
        }
        return render(request, "admin/add_tasks_to_shot.html", context=context)

    @admin.action(description="Добавить комментарий")
    def add_comments_to_shot(modeladmin, request, queryset):
        if "apply" in request.POST:
            form = AddСommentToShotForm(request.POST)
            if form.is_valid():
                comment = form.cleaned_data["comment"]
                created_by = form.cleaned_data["created_by"]
                for shot in queryset:
                    ChatMessage.objects.create(
                        shot=shot,
                        text=comment,
                        created_by=created_by,
                        created_at=datetime.now(),
                    )

            return HttpResponseRedirect(request.get_full_path())

        context = {
            "shots": queryset,
            "form": AddСommentToShotForm,
        }
        return render(
            request,
            "admin/add_comments_to_shot.html",
            context=context,
        )

    @admin.action(description="Добавить в группы")
    def add_shots_to_groups(modeladmin, request, queryset):
        if "apply" in request.POST:
            form = AddShotsToGroupsForm(request.POST)
            if form.is_valid():
                groups = form.cleaned_data["groups"]
                for shot in queryset:
                    shot.group.set(shot.group.all().union(groups))
                    shot.save()

            return HttpResponseRedirect(request.get_full_path())

        context = {
            "shots": queryset,
            "form": AddShotsToGroupsForm,
        }
        return render(request, "admin/add_shots_to_groups.html", context=context)

    @admin.action(description="Удалить из групп")
    def remove_shots_from_groups(modeladmin, request, queryset):
        if "apply" in request.POST:
            form = AddShotsToGroupsForm(request.POST)
            if form.is_valid():
                groups = form.cleaned_data["groups"]
                for shot in queryset:
                    shot.group.remove(*groups)
                    shot.save()

            return HttpResponseRedirect(request.get_full_path())

        context = {
            "shots": queryset,
            "form": AddShotsToGroupsForm,
        }
        return render(request, "admin/remove_shots_from_groups.html", context=context)

    @admin.action(description="Скачать в виде Excel таблицы")
    def download_shots_as_xlsx(modeladmin, request, queryset):
        from datetime import datetime

        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        project_code = queryset[0].group.first().project.code
        date = datetime.today().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"{project_code}_shots_{date}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"

        wb = Workbook()
        ws = wb.active
        ws.append(
            (
                "№",
                "Название шота",
                "Таймкод",
                "Задачи",
            )
        )
        for counter, shot in enumerate(queryset, start=1):
            counter_cell = ws.cell(row=counter + 1, column=1, value=counter)
            counter_cell.alignment = Alignment(vertical="top", horizontal="left")

            name_cell = ws.cell(row=counter + 1, column=2, value=shot.name)
            name_cell.alignment = Alignment(vertical="top", horizontal="left")

            name_cell = ws.cell(row=counter + 1, column=3, value=shot.rec_timecode)
            name_cell.alignment = Alignment(vertical="top", horizontal="left")

            ws.cell(
                row=counter + 1,
                column=4,
                value="\n".join([task.description for task in shot.task.all()]),
            )

        for i, column in enumerate(ws.columns, 1):
            lengths = (
                len(max(str(cell.value).split("\n"))) for cell in column if cell.value is not None
            )
            ws.column_dimensions[get_column_letter(i)].width = max(lengths) + 2

        wb.save(response)
        return response

    def save_model(self, request, obj, form, change):
        obj.created_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(Version)
class VersionAdmin(admin.ModelAdmin):
    save_on_top = True
    readonly_fields = (
        "created_by",
        "created_at",
    )
    list_display = (
        "name",
        "created_at",
        "created_by",
    )
    ordering = ("name",)

    def save_model(self, request, obj, form, change):
        obj.created_by = request.user
        super().save_model(request, obj, form, change)
        preview_path = f"{obj.video.path}.jpg"
        cmd = ["ffmpeg", "-y", "-i", obj.video.path, "-frames:v", "1", preview_path]
        subprocess.run(cmd)
        obj.preview.save(Path(preview_path).name, File(open(preview_path, "rb")))


@admin.register(Task)
class TaskAdmin(admin.ModelAdmin):
    save_on_top = True
    list_display = (
        "description",
        "default_status",
        "project",
        "created_at",
    )
    readonly_fields = (
        "created_by",
        "created_at",
    )
    list_filter = ("project",)
    ordering = ("description",)
    inlines = (ShotTaskInline,)

    actions = ["add_shots_to_project"]

    @admin.action(description="Добавить задачи в проект")
    def add_shots_to_project(modeladmin, request, queryset):
        if "apply" in request.POST:
            form = AddShotsToProjectForm(request.POST)
            if form.is_valid():
                project = form.cleaned_data["project"]
                for task in queryset:
                    task.project = project
                    task.save()

            return HttpResponseRedirect(request.get_full_path())

        context = {
            "shots": queryset,
            "form": AddShotsToProjectForm,
        }
        return render(request, "admin/add_shots_to_project.html", context=context)

    def save_model(self, request, obj, form, change):
        obj.created_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(Status)
class StatusAdmin(admin.ModelAdmin):
    save_on_top = True
    list_display = (
        "title",
        "get_color",
    )
    readonly_fields = (
        "created_by",
        "created_at",
    )

    @admin.display(description="цвет")
    def get_color(self, obj):
        return mark_safe(f"<div style='background-color:{obj.color};color:#fff'>{obj.color}</div>")

    def save_model(self, request, obj, form, change):
        obj.created_by = request.user
        super().save_model(request, obj, form, change)


@admin.register(ShotTask)
class ShotTaskAdmin(admin.ModelAdmin):
    save_on_top = True
    list_display = (
        "shot",
        "task",
        "hours",
        "status",
        "assigned_to",
    )


@admin.register(TmpShotPreview)
class TmpShotPreview(admin.ModelAdmin):
    save_on_top = True


@admin.register(ChatMessage)
class ChatMessageAdmin(admin.ModelAdmin):
    save_on_top = True


@admin.register(Attachment)
class AttachmentAdmin(admin.ModelAdmin):
    save_on_top = True
